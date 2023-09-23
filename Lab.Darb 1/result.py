from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
max_row=ws.max_row
for row in range(2,max_row+1):
    hours=ws['B' + str(row)].value
    rate=ws['C' + str(row)].value
    if(type(hours)!=str and type(rate)!=str):
        salary=hours*rate
        ws['D'+str(row)].value=salary
        print(salary)
        if(salary>=3000):
            total+=1
    else:
        continue
wb.save('result.xlsx')

print("Cilvēkiem kam alga ir lielāka par 3000: "+str(total))
wb.close()
