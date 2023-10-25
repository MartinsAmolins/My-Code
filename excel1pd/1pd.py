
###Darbs ar excel jeb xslx fialiem

from openpyxl import Workbook, load_workbook 

wb=load_workbook('EmployeeData.xlsx')
ws=wb.active
total=0
max_row=ws.max_row
data=[]
for row in range(2,max_row+1):
    full_name=ws['A' + str(row)].value
    jobt_title=ws['B' + str(row)].value
    department=ws['C' + str(row)].value
    gender=ws['D' + str(row)].value
    age=ws['E' + str(row)].value
    annual_salary=ws['F' + str(row)].value
    bonus=ws['G' + str(row)].value
    country=ws['H' + str(row)].value
data.append(full_name)
data.append(jobt_title)
data.append(department)
data.append(gender)


wb.save('result.xlsx')


wb.close()

