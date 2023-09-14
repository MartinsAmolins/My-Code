from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
max_row=ws.max_row
for row in range(2,max_row+1):
    hours=ws['B' + str(row)].value
    rate=ws['C' + str(row)].value
    if(type(hours)==int and type(rate)==int):
        salary=hours*rate
        ws['D'+str(row)].value=salary
        print(salary)
    else:
        continue
wb.save('result.xlsx')

print(total)
wb.close()
